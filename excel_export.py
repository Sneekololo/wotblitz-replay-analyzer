import tempfile

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


KEYS = [
    "account_id", "nickname", "battles",
    "HT", "MT", "LT", "TD", "main_tank",
    "ADR", "frags", "KPR", "DE", "assist", "blocked",
    "shots", "hits", "pens",
    "AccH", "AccP",
    "iPoints", "sPoints",
    "Firepower", "AIM", "Support", "Supremacy", "BPR",
]

HEADERS = [
    "account_id", "nickname", "battles",
    "HT", "MT", "LT", "TD", "main_tank",
    "ADR", "Frags", "KPR", "DE", "Assist", "Blocked",
    "shots", "hits", "pens",
    "AccH (%)", "AccP (%)",
    "iPoints", "sPoints",
    "Firepower", "AIM", "Support", "Supremacy", "BPR 2.0",
]

COL_WIDTHS = [14, 22, 9, 6, 6, 6, 6, 18, 8, 7, 7, 7, 9, 9, 8, 8, 8, 10, 10, 9, 9, 11, 9, 10, 11, 9]


def send_stats_workbook(our_rows, enemy_rows):
    workbook = Workbook()
    write_sheet(workbook.active, "Our Team", our_rows, "FF5500")
    write_sheet(workbook.create_sheet("Enemy Team"), "Enemy Team", enemy_rows, "444444")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    workbook.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def write_sheet(sheet, title, rows, title_color):
    sheet.title = title
    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor=title_color)
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    even_fill = PatternFill("solid", fgColor="1A1A1A")
    odd_fill = PatternFill("solid", fgColor="141414")
    data_font = Font(color="E0E0E0", name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center")
    bottom_border = Border(bottom=Side(style="thin", color="2A2A2A"))

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for index, player in enumerate(rows):
        sheet.append([player.get(key, "") for key in KEYS])
        fill = even_fill if index % 2 == 0 else odd_fill
        for cell in sheet[index + 2]:
            cell.fill = fill
            cell.font = data_font
            cell.alignment = center
            cell.border = bottom_border

    for index, width in enumerate(COL_WIDTHS, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.row_dimensions[1].height = 22
