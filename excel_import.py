from openpyxl import load_workbook

from excel_export import HEADERS, KEYS


def normalize_header(value):
    return "".join(ch.lower() if ch.isalnum() else " " for ch in str(value or "")).strip()


HEADER_TO_KEY = {
    normalize_header(header): key
    for header, key in zip(HEADERS, KEYS)
}
HEADER_TO_KEY.update({
    "player": "nickname",
    "name": "nickname",
    "tank": "main_tank",
    "main tank": "main_tank",
    "bpr": "BPR",
    "bpr 2": "BPR",
    "bpr 20": "BPR",
    "accuracy": "AccH",
    "hit rate": "AccH",
    "pen rate": "AccP",
})


def parse_team_database(file_obj, max_rows=1000):
    workbook = load_workbook(file_obj, data_only=True, read_only=True)
    players = []

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            continue

        keys = [HEADER_TO_KEY.get(normalize_header(header)) for header in headers]
        if "nickname" not in keys:
            continue

        for row in rows:
            if len(players) >= max_rows:
                break
            player = {"source_sheet": sheet.title}
            for index, value in enumerate(row):
                if index >= len(keys) or not keys[index]:
                    continue
                player[keys[index]] = clean_cell(value)
            if player.get("nickname"):
                players.append(player)

    return players


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value
